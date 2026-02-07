"""Export service for dialogue export to Excel."""
import io
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class ExportService:
    """Service for exporting Q&A dialogues to various formats."""
    
    @staticmethod
    def create_dialogue_excel(
        dialogues: list[dict],
        username: str,
        session_summary: Optional[dict] = None
    ) -> bytes:
        """
        Create an Excel file from Q&A dialogues.
        
        Args:
            dialogues: List of Q&A pairs with format:
                [{"question": str, "answer": str, "evidence": dict, "timestamp": str}, ...]
            username: Username for the export
            session_summary: Optional session context (layer summary, etc.)
        
        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        
        # ===== Q&A Sheet =====
        ws_qa = wb.active
        ws_qa.title = "Q&A Dialogue"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="2F2F2F", end_color="2F2F2F", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='E2E2E2'),
            right=Side(style='thin', color='E2E2E2'),
            top=Side(style='thin', color='E2E2E2'),
            bottom=Side(style='thin', color='E2E2E2')
        )
        
        # Alternating row colors
        row_fill_even = PatternFill(start_color="F6F6F6", end_color="F6F6F6", fill_type="solid")
        
        # Headers
        headers = ["#", "Question", "Answer", "Document Evidence", "Session Context", "Timestamp"]
        for col, header in enumerate(headers, 1):
            cell = ws_qa.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows
        for idx, dialogue in enumerate(dialogues, 1):
            row = idx + 1
            
            # Format evidence
            evidence_text = ""
            if dialogue.get("evidence"):
                ev = dialogue["evidence"]
                if ev.get("document_chunks"):
                    chunks = ev["document_chunks"][:5]  # Max 5 chunks
                    evidence_text = "\n".join([
                        f"- {c.get('source', 'Unknown')} (p{c.get('page', '?')}) - {c.get('section', 'general')}"
                        for c in chunks
                    ])
            
            # Format session context
            session_text = ""
            if dialogue.get("evidence", {}).get("session_objects"):
                so = dialogue["evidence"]["session_objects"]
                if so.get("layers_used"):
                    session_text = f"Layers: {', '.join(so['layers_used'])}"
                if so.get("objects_count"):
                    session_text += f"\nObjects: {so['objects_count']}"
            
            # Write cells
            cells_data = [
                idx,
                dialogue.get("question", ""),
                dialogue.get("answer", ""),
                evidence_text,
                session_text,
                dialogue.get("timestamp", "")
            ]
            
            for col, value in enumerate(cells_data, 1):
                cell = ws_qa.cell(row=row, column=col, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
                if row % 2 == 0:
                    cell.fill = row_fill_even
        
        # Column widths
        column_widths = [5, 40, 60, 35, 25, 18]
        for col, width in enumerate(column_widths, 1):
            ws_qa.column_dimensions[get_column_letter(col)].width = width
        
        # Row heights
        ws_qa.row_dimensions[1].height = 25
        for row in range(2, len(dialogues) + 2):
            ws_qa.row_dimensions[row].height = 80
        
        # ===== Summary Sheet =====
        if session_summary or dialogues:
            ws_summary = wb.create_sheet("Summary")
            
            # Title
            ws_summary.cell(row=1, column=1, value="AICI Q&A Export Summary")
            ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)
            ws_summary.merge_cells('A1:C1')
            
            # Export info
            info_data = [
                ("Export Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                ("User:", username),
                ("Total Questions:", len(dialogues)),
            ]
            
            if session_summary:
                if session_summary.get("object_count"):
                    info_data.append(("Drawing Objects:", session_summary["object_count"]))
                if session_summary.get("layer_summary"):
                    layers = session_summary["layer_summary"]
                    info_data.append(("Layers:", ", ".join(f"{k} ({v})" for k, v in layers.items())))
            
            for row, (label, value) in enumerate(info_data, 3):
                ws_summary.cell(row=row, column=1, value=label).font = Font(bold=True)
                ws_summary.cell(row=row, column=2, value=value)
            
            ws_summary.column_dimensions['A'].width = 20
            ws_summary.column_dimensions['B'].width = 50
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


# Singleton instance
_export_service: Optional[ExportService] = None


def get_export_service() -> ExportService:
    """Get or create export service instance."""
    global _export_service
    if _export_service is None:
        _export_service = ExportService()
    return _export_service
